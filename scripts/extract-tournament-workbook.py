import argparse
import json
from pathlib import Path

from openpyxl import load_workbook


TARGET_SHEETS = ("赛事对战表", "手游横版id名单", "端游横版id名单")


def cell_value(cell):
    value = cell.value
    if value is None:
        return None
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return value


def extract_sheet(sheet):
    # Some WPS exports incorrectly declare every worksheet dimension as A1.
    # Read-only openpyxl trusts that cache unless dimensions are reset first.
    sheet.reset_dimensions()
    rows = []
    for row in sheet.iter_rows():
        values = [cell_value(cell) for cell in row]
        if any(value not in (None, "") for value in values):
            rows.append({
                "row": row[0].row,
                "values": values,
            })
    return {
        "maxRow": sheet.max_row,
        "maxColumn": sheet.max_column,
        "rows": rows,
    }


parser = argparse.ArgumentParser()
parser.add_argument("workbook")
parser.add_argument("output")
args = parser.parse_args()

workbook_path = Path(args.workbook).resolve()
workbook = load_workbook(workbook_path, read_only=True, data_only=True)
result = {
    "workbook": str(workbook_path),
    "sheetNames": workbook.sheetnames,
    "sheets": {},
}
for name in TARGET_SHEETS:
    if name in workbook.sheetnames:
        result["sheets"][name] = extract_sheet(workbook[name])
workbook.close()

output_path = Path(args.output).resolve()
output_path.parent.mkdir(parents=True, exist_ok=True)
output_path.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
print(output_path)
